import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from adopters.models import Adopter
from appointments.models import Appointment
from bookings.models import Booking
from pending_adoptions.enums import PendingAdoptionStatus
from pending_adoptions.models import PendingAdoption
from utils import DateTimeUtils
from users.models import UserProfile


class ContinuityAccessSpreadsheetService:
    def __init__(self, date=DateTimeUtils.get_today()):
        self.wb = Workbook()
        self.date = date

    def write_headers(self, ws, headers: list[str]):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top")

    def create_schedule_sheet(self):
        """Create the Schedule sheet with appointment data"""
        wb = self.wb
        ws = wb.create_sheet("Schedule", 0)

        # Define headers
        headers = [
            "Time",
            "Appointment",
            "Email",
            "Phone",
            "Booked Instant",
            "Notes from Adopter",
            "Notes from Adoptions",
            "Bringing Dog",
            "Has Cats",
            "Mobility?",
        ]
        self.write_headers(ws, headers)

        appts = Appointment.objects.filter(
            instant__range=DateTimeUtils.get_range_for_date(self.date),
            soft_deleted=False,
        ).order_by("instant")

        schedule_data = []

        for appt in appts:
            booking: Booking = appt.get_current_booking()
            if booking is None and not appt.is_admin_appointment:
                continue

            if appt.is_adoption_appointment:
                adopter: Adopter = booking.adopter
                user: UserProfile = adopter.user_profile

                one_appt = [
                    appt.time_display,
                    f"{appt.description} ({appt.type_display})",
                    adopter.primary_email,
                    user.phone_number,
                    booking.created_display,
                    adopter.adopter_notes,
                    adopter.internal_notes,
                    ("Yes" if adopter.bringing_dog else ""),
                    ("Yes" if adopter.cats_in_home else ""),
                    ("Yes" if adopter.mobility else ""),
                ]
            elif appt.is_admin_appointment:
                one_appt = [
                    appt.time_display,
                    f"{appt.description} ({appt.type_display})",
                    "",
                    "",
                    "",
                    "",
                    appt.appointment_notes,
                ]

            schedule_data.append(one_appt)

        if len(schedule_data) == 0:
            cell = ws.cell(row=2, column=1, value="No appointments today!")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
            return

        # Write data rows
        for row_num, row_data in enumerate(schedule_data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

        # Set column widths
        column_widths = {
            "A": 12,  # Time
            "B": 25,  # Appointment Description/Type
            "C": 25,  # Email
            "D": 18,  # Phone
            "E": 22,  # Booked Instant
            "F": 25,  # Notes from Adopter
            "G": 25,  # Notes from Adoptions
            "H": 15,  # Bringing Dog
            "I": 12,  # Has Cats
            "J": 12,  # Mobility?
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def create_chosen_board_sheet(self):
        """Create the Chosen Board sheet with pending adoption data"""
        wb = self.wb
        ws = wb.create_sheet("Chosen Board", 1)

        # Define headers
        headers = ["Dog", "Adopter", "Email", "Status", "HW?"]
        self.write_headers(ws, headers)

        adoptions = (
            PendingAdoption.objects.filter(
                paperwork_appointment=None,
            )
            .exclude(
                status=PendingAdoptionStatus.CANCELED,
            )
            .exclude(
                status=PendingAdoptionStatus.COMPLETED,
            )
            .order_by("-status", "dog")
        )

        chosen_board_data = []

        for adoption in adoptions:            
            one_adoption = [
                adoption.dog,
                adoption.adopter.user_profile.full_name,
                adoption.adopter.primary_email,
                adoption.status_display,
                adoption.hw_display
            ]

            chosen_board_data.append(one_adoption)

        if len(chosen_board_data) == 0:
            cell = ws.cell(row=2, column=1, value="No adoptions on the Chosen Board!")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
            return

        # Write data rows
        for row_num, row_data in enumerate(chosen_board_data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top")

        # Set column widths
        column_widths = {
            "A": 18,  # Dog
            "B": 22,  # Adopter
            "C": 22,  # Email
            "D": 20,  # Status
            "E": 12,  # HW?
            "F": 22,  # Last Update Sent
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def create_schedule_export(self):
        """
        Create an Excel workbook with Schedule and Chosen Board sheets.
        Hardcoded with sample data matching the provided template.
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create Schedule sheet
        self.create_schedule_sheet()

        # Create Chosen Board sheet
        self.create_chosen_board_sheet()

        # Save to BytesIO buffer
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)

        return buffer

    if __name__ == "__main__":
        output_file = "/mnt/user-data/outputs/schedule_export_generated.xlsx"
        create_schedule_export(output_file)


class ReportingStatsExportService:
    _PERIOD_LABELS = {
        "ytd": "YTD",
        "pytd": "PYTD",
        "prev_year": "Prev Year",
        "current_month": "Curr Month",
        "current_month_py": "Curr Month PY",
    }
    _PERIOD_KEYS = list(_PERIOD_LABELS.keys())

    def __init__(self, periods: dict, global_stats: dict):
        self.periods = periods
        self.global_stats = global_stats
        self.wb = Workbook()

    def _write_headers(self, ws, headers: list[str]) -> None:
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top")

    def _write_rows(self, ws, rows: list[list], start_row: int = 2) -> None:
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top")

    def _outcome_rows(self) -> list[list]:
        metrics = [
            ("Total Appointments", "total"),
            ("Adoptions", "adoptions"),
            ("Chosen", "chosen"),
            ("FTA", "fta"),
            ("No Decision", "no_decision"),
            ("No Show", "no_show"),
            ("HW+", "hw_positive"),
            ("HW-", "hw_negative"),
        ]
        rows = []
        for label, key in metrics:
            ytd = self.periods["ytd"][key]
            pytd = self.periods["pytd"][key]
            delta = ytd - pytd
            delta_pct = round((delta / pytd * 100), 1) if pytd else None
            row = [label] + [self.periods[p][key] for p in self._PERIOD_KEYS]
            row.append(f"+{delta}" if delta > 0 else str(delta))
            row.append(f"{delta_pct:+.1f}%" if delta_pct is not None else "—")
            rows.append(row)
        return rows

    def _pipeline_rows(self) -> list[list]:
        metrics = [
            ("Applications Processed", "applications"),
            ("Visitors", "visitors"),
            ("Adopted", "adopters"),
        ]
        rows = []
        for label, key in metrics:
            row = [label] + [self.periods[p][key] for p in self._PERIOD_KEYS]
            rows.append(row)
        return rows

    def _create_outcomes_sheet(self) -> None:
        ws = self.wb.create_sheet("Appointment Outcomes", 0)
        headers = ["Metric"] + list(self._PERIOD_LABELS.values()) + ["Δ (YTD vs PYTD)", "Δ%"]
        self._write_headers(ws, headers)
        self._write_rows(ws, self._outcome_rows())
        ws.column_dimensions["A"].width = 22

    def _create_pipeline_sheet(self) -> None:
        ws = self.wb.create_sheet("Adopter Pipeline", 1)
        headers = ["Metric"] + list(self._PERIOD_LABELS.values())
        self._write_headers(ws, headers)
        self._write_rows(ws, self._pipeline_rows())
        ws.column_dimensions["A"].width = 26

    def _create_global_sheet(self) -> None:
        ws = self.wb.create_sheet("Global Stats", 2)
        self._write_headers(ws, ["Metric", "Value"])
        rows = [
            ["Avg visits before adopting", round(self.global_stats["avg_visits_before_adopting"], 1)],
            ["Visitors who never adopted (all-time)", self.global_stats["visitors_never_adopted"]],
        ]
        self._write_rows(ws, rows)
        ws.column_dimensions["A"].width = 36

    def build(self) -> io.BytesIO:
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self._create_outcomes_sheet()
        self._create_pipeline_sheet()
        self._create_global_sheet()
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
